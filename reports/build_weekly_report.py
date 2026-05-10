"""
reports/build_weekly_report.py — Weekly Excel of all backtest output.

Re-runs the canonical 1-lot backtest, then assembles `reports/weekly_backtest.xlsx`
with these tabs:

    SUMMARY                — one-row-per-strategy stats across the full window
    VP_TRAIL_MONTHLY       — monthly P&L / WR / trade count for VP-Trail-Swing
    VP_TRAIL_TRADES        — every VP-Trail trade
    ORB_MONTHLY            — same for futures ORB
    ORB_TRADES
    OPT_ORB_MONTHLY        — options ORB
    OPT_ORB_TRADES
    VWAP_MONTHLY           — VWAP_REV (archived but tracked)
    VWAP_TRADES
    CREDIT_SPREAD_MONTHLY  — credit spread (archived but tracked)
    CREDIT_SPREAD_TRADES

Run:
    python -m reports.build_weekly_report
    python -m reports.build_weekly_report --skip-rerun     # use existing logs
    python -m reports.build_weekly_report --out path.xlsx
"""

from __future__ import annotations

import argparse
import glob
import os
import pathlib
import subprocess
import sys
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

TRADE_DIR = ROOT / 'trade_logs'
OUT_DIR   = ROOT / 'reports'

ACTIVE = {'VP_TRAIL', 'ORB', 'OPT_ORB'}   # currently deployed
ARCHIVED = {'VWAP_REV', 'CREDIT_SPREAD'}   # tracked for reference

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUB_FILL    = PatternFill('solid', fgColor='D9E1F2')
WIN_FILL    = PatternFill('solid', fgColor='C6EFCE')
LOSS_FILL   = PatternFill('solid', fgColor='FFC7CE')


# ─────────────────────────────────────────────────────────────────────────────
# Loaders
# ─────────────────────────────────────────────────────────────────────────────
def load_vp_trail() -> pd.DataFrame:
    parts = []
    for f in sorted(glob.glob(str(TRADE_DIR / 'vpt_final_*_*S.csv'))):
        name = pathlib.Path(f).stem
        # vpt_final_BANKNIFTY_IS / _OOS
        bits = name.split('_')
        inst, tag = bits[2], bits[3]
        d = pd.read_csv(f)
        d['instrument'] = inst
        d['period']     = tag
        d['strategy']   = 'VP_TRAIL'
        parts.append(d)
    if not parts:
        return pd.DataFrame()
    df = pd.concat(parts, ignore_index=True)
    df['entry_ts'] = pd.to_datetime(df['entry_ts'])
    df['exit_ts']  = pd.to_datetime(df.get('exit_ts'), errors='coerce')
    return df


def load_full_backtest() -> pd.DataFrame:
    """ORB / VWAP_REV / OPT_ORB live in full_backtest_*.csv."""
    parts = []
    for f in sorted(glob.glob(str(TRADE_DIR / 'full_backtest_*.csv'))):
        parts.append(pd.read_csv(f))
    if not parts:
        return pd.DataFrame()
    df = pd.concat(parts, ignore_index=True)
    df = df.drop_duplicates(['trade_id', 'strategy', 'date', 'entry_time'],
                            keep='last')
    df['entry_ts'] = pd.to_datetime(df['entry_time'])
    df['exit_ts']  = pd.to_datetime(df.get('exit_time'), errors='coerce')
    return df


def load_credit_spread() -> pd.DataFrame:
    parts = []
    for f in sorted(glob.glob(str(TRADE_DIR / 'spr_swing_ext_*.csv'))):
        d = pd.read_csv(f)
        d['strategy'] = 'CREDIT_SPREAD'
        parts.append(d)
    if not parts:
        return pd.DataFrame()
    df = pd.concat(parts, ignore_index=True)
    df['entry_ts'] = pd.to_datetime(df['entry_ts'])
    df['exit_ts']  = pd.to_datetime(df.get('exit_ts'), errors='coerce')
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Stats helpers
# ─────────────────────────────────────────────────────────────────────────────
def _stats_row(df: pd.DataFrame, label: str) -> dict:
    if df.empty:
        return {'label': label, 'n': 0, 'wr_pct': 0.0, 'pnl_rs': 0.0,
                'avg_pnl': 0.0, 'max_win': 0.0, 'max_loss': 0.0,
                'max_dd_rs': 0.0, 'first_date': None, 'last_date': None}
    pnl = df['pnl_rs'].astype(float)
    wins = (pnl > 0).astype(int)
    cum = pnl.sort_index().cumsum().values
    dd = float((np.maximum.accumulate(cum) - cum).max()) if len(cum) else 0
    return {
        'label':     label,
        'n':         int(len(df)),
        'wr_pct':    round(wins.mean() * 100, 1),
        'pnl_rs':    round(float(pnl.sum()), 0),
        'avg_pnl':   round(float(pnl.mean()), 0),
        'max_win':   round(float(pnl.max()), 0),
        'max_loss':  round(float(pnl.min()), 0),
        'max_dd_rs': round(dd, 0),
        'first_date': df['entry_ts'].min(),
        'last_date':  df['entry_ts'].max(),
    }


def monthly_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=['month','n','wr_pct','pnl_rs','avg_pnl','max_loss'])
    df = df.copy()
    df['month'] = df['entry_ts'].dt.to_period('M').astype(str)
    g = df.groupby('month').agg(
        n      =('pnl_rs', 'count'),
        wins   =('pnl_rs', lambda s: (s > 0).sum()),
        pnl_rs =('pnl_rs', 'sum'),
        max_loss=('pnl_rs', 'min'),
    ).reset_index()
    g['wr_pct'] = (g['wins'] / g['n'] * 100).round(1)
    g['avg_pnl'] = (g['pnl_rs'] / g['n']).round(0)
    g['pnl_rs']  = g['pnl_rs'].round(0)
    g['max_loss'] = g['max_loss'].round(0)
    return g[['month', 'n', 'wr_pct', 'pnl_rs', 'avg_pnl', 'max_loss']]


# ─────────────────────────────────────────────────────────────────────────────
# Sheet writers
# ─────────────────────────────────────────────────────────────────────────────
def _write_header(ws, headers: list[str], row: int = 1):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(h)) + 2)


def _autosize(ws, df: pd.DataFrame):
    for i, col in enumerate(df.columns, start=1):
        try:
            sample = df[col].astype(str).str.len().clip(upper=40).max()
        except Exception:
            sample = 12
        width = max(12, int(sample) + 2 if sample else 12, len(col) + 2)
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)


def write_summary_sheet(wb: Workbook, datasets: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet('SUMMARY', 0)
    ws['A1'] = 'Hawala v2 — Weekly backtest summary'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Generated: {datetime.now():%Y-%m-%d %H:%M}'
    ws['A2'].font = Font(italic=True, color='606060')

    headers = ['Strategy', 'Status', 'Trades', 'WR %', 'Net P&L (₹)',
               'Avg/trade (₹)', 'Max win (₹)', 'Max loss (₹)',
               'Max DD (₹)', 'First trade', 'Last trade']
    _write_header(ws, headers, row=4)

    r = 5
    for strat, df in datasets.items():
        s = _stats_row(df, strat)
        status = 'ACTIVE' if strat in ACTIVE else 'ARCHIVED'
        ws.cell(row=r, column=1, value=strat).font = Font(bold=True)
        ws.cell(row=r, column=2, value=status)
        ws.cell(row=r, column=3, value=s['n'])
        ws.cell(row=r, column=4, value=s['wr_pct'])
        c = ws.cell(row=r, column=5, value=s['pnl_rs'])
        c.fill = WIN_FILL if s['pnl_rs'] >= 0 else LOSS_FILL
        ws.cell(row=r, column=6, value=s['avg_pnl'])
        ws.cell(row=r, column=7, value=s['max_win'])
        ws.cell(row=r, column=8, value=s['max_loss'])
        ws.cell(row=r, column=9, value=s['max_dd_rs'])
        ws.cell(row=r, column=10, value=str(s['first_date'])[:10] if s['first_date'] else '')
        ws.cell(row=r, column=11, value=str(s['last_date'])[:10] if s['last_date'] else '')
        if status == 'ARCHIVED':
            for col in range(1, 12):
                ws.cell(row=r, column=col).font = Font(italic=True, color='808080')
        r += 1

    # Combined active row
    active_df = pd.concat([df for s, df in datasets.items()
                           if s in ACTIVE and not df.empty], ignore_index=True) \
        if any(s in ACTIVE for s in datasets) else pd.DataFrame()
    s = _stats_row(active_df, 'ALL ACTIVE')
    ws.cell(row=r, column=1, value='ALL ACTIVE').font = Font(bold=True, color='1F4E79')
    ws.cell(row=r, column=2, value='—')
    ws.cell(row=r, column=3, value=s['n'])
    ws.cell(row=r, column=4, value=s['wr_pct'])
    c = ws.cell(row=r, column=5, value=s['pnl_rs'])
    c.fill = WIN_FILL if s['pnl_rs'] >= 0 else LOSS_FILL
    c.font = Font(bold=True)
    ws.cell(row=r, column=6, value=s['avg_pnl'])
    ws.cell(row=r, column=7, value=s['max_win'])
    ws.cell(row=r, column=8, value=s['max_loss'])
    ws.cell(row=r, column=9, value=s['max_dd_rs'])
    for col in range(1, 12):
        ws.cell(row=r, column=col).fill = SUB_FILL if col != 5 else WIN_FILL if s['pnl_rs'] >= 0 else LOSS_FILL

    ws.freeze_panes = 'A5'


def write_monthly_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    ws['A1'] = f'{name} — monthly breakdown'
    ws['A1'].font = Font(bold=True, size=12)

    monthly = monthly_breakdown(df)
    if monthly.empty:
        ws['A3'] = '(no trades)'
        return

    headers = list(monthly.columns)
    _write_header(ws, headers, row=3)

    for i, row in monthly.iterrows():
        r = 4 + i
        for j, col in enumerate(headers, start=1):
            v = row[col]
            cell = ws.cell(row=r, column=j, value=v)
            if col == 'pnl_rs':
                cell.fill = WIN_FILL if v >= 0 else LOSS_FILL
            elif col == 'wr_pct':
                if v >= 50:   cell.fill = WIN_FILL
                elif v <= 30: cell.fill = LOSS_FILL

    # Totals row
    r = 4 + len(monthly)
    ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=r, column=2, value=int(monthly['n'].sum())).font = Font(bold=True)
    ws.cell(row=r, column=3, value=round(
        (df['pnl_rs'] > 0).mean() * 100, 1) if not df.empty else 0).font = Font(bold=True)
    pnl = float(df['pnl_rs'].sum()) if not df.empty else 0
    c = ws.cell(row=r, column=4, value=round(pnl, 0))
    c.font = Font(bold=True)
    c.fill = WIN_FILL if pnl >= 0 else LOSS_FILL

    ws.freeze_panes = 'A4'
    _autosize(ws, monthly)


def write_trades_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    if df.empty:
        ws['A1'] = '(no trades)'
        return
    cols_pref = ['entry_ts', 'exit_ts', 'instrument', 'strategy', 'direction',
                 'entry', 'exit_price', 'stop', 'target',
                 'pnl_pts', 'pnl_rs', 'win', 'exit_reason',
                 'period', 'contract', 'atr14', 'vah', 'val', 'poc']
    cols = [c for c in cols_pref if c in df.columns]
    out = df[cols].sort_values('entry_ts').copy()
    if 'entry_ts' in out: out['entry_ts'] = out['entry_ts'].astype(str)
    if 'exit_ts'  in out: out['exit_ts']  = out['exit_ts'].astype(str)

    _write_header(ws, cols, row=1)
    for i, row in out.iterrows():
        for j, col in enumerate(cols, start=1):
            v = row[col]
            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
                v = None
            cell = ws.cell(row=i + 2, column=j, value=v)
            if col == 'pnl_rs' and isinstance(v, (int, float)):
                cell.fill = WIN_FILL if v >= 0 else LOSS_FILL

    ws.freeze_panes = 'A2'
    _autosize(ws, out)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def maybe_rerun_canonical():
    print('  ▶ refreshing vpt_final_*.csv via run_canonical.py …')
    r = subprocess.run(
        [sys.executable, str(ROOT / 'run_canonical.py')],
        cwd=str(ROOT), capture_output=True, text=True, timeout=600,
    )
    if r.returncode != 0:
        print('  ⚠ run_canonical failed:')
        print(r.stderr[-2000:])
    else:
        # surface the COMBINED block for the operator
        for line in r.stdout.splitlines()[-8:]:
            print('    ' + line)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--skip-rerun', action='store_true',
                    help='skip the canonical re-run; use existing CSVs')
    ap.add_argument('--out', default=None)
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if not args.skip_rerun:
        maybe_rerun_canonical()

    # ── Load all strategies ────────────────────────────────────────────────
    vp = load_vp_trail()
    fb = load_full_backtest()
    sp = load_credit_spread()

    datasets: dict[str, pd.DataFrame] = {
        'VP_TRAIL':      vp,
        'ORB':           fb[fb['strategy'] == 'ORB']      if not fb.empty else fb,
        'OPT_ORB':       fb[fb['strategy'] == 'OPT_ORB']  if not fb.empty else fb,
        'VWAP_REV':      fb[fb['strategy'] == 'VWAP_REV'] if not fb.empty else fb,
        'CREDIT_SPREAD': sp,
    }

    # ── Build workbook ─────────────────────────────────────────────────────
    wb = Workbook()
    # Remove the default sheet — we add SUMMARY at index 0 manually
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    write_summary_sheet(wb, datasets)
    for strat, df in datasets.items():
        write_monthly_sheet(wb, f'{strat}_MONTHLY', df)
        write_trades_sheet(wb, f'{strat}_TRADES', df)

    out_path = pathlib.Path(args.out) if args.out \
        else OUT_DIR / 'weekly_backtest.xlsx'
    wb.save(out_path)

    size_kb = out_path.stat().st_size / 1024
    n_total = sum(len(df) for df in datasets.values())
    print(f'  ✓ wrote {out_path}  ({size_kb:.1f} KB, '
          f'{n_total} trades across {len(datasets)} strategies)')


if __name__ == '__main__':
    main()
