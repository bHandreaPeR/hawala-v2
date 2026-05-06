"""
v3/scripts/update_excel_nifty.py
==================================
Reads trade_log_options_nifty.csv + trade_log_options_banknifty.csv
and updates Hawala_v3_Trade_Summary.xlsx:
  - Nifty v3 Trades sheet
  - BankNifty v3 Trades sheet
  - Summary sheet

Usage:
    cd "Hawala v2/Hawala v2"
    python3 v3/scripts/update_excel_nifty.py
"""
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT     = Path(__file__).resolve().parents[2]
CSV_N    = ROOT / 'trade_log_options_nifty.csv'
CSV_BN   = ROOT / 'trade_log_options_banknifty.csv'
XLSX     = ROOT.parent / 'Hawala_v3_Trade_Summary.xlsx'

LOT_N    = 65     # Nifty lot size
LOT_BN   = 30    # BankNifty lot size

# ── colours ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = '1F3864'
C_HEADER_FG  = 'FFFFFF'
C_WIN_BG     = 'E2EFDA'
C_LOSS_BG    = 'FCE4D6'
C_TOTAL_BG   = 'D9E1F2'

def _thin():
    s = Side(border_style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=C_HEADER_FG, name='Arial', size=10)
    c.fill      = PatternFill('solid', start_color=C_HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = _thin()
    if width and ws.column_dimensions[get_column_letter(col)].width < width:
        ws.column_dimensions[get_column_letter(col)].width = width

def _data_cell(ws, row, col, value, bold=False, fmt=None, bg=None, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, name='Arial', size=9)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border    = _thin()
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _clear_sheet(ws):
    """Unmerge then clear — avoids MergedCell read-only AttributeError."""
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value  = None
                cell.fill   = PatternFill(fill_type=None)
                cell.font   = Font()
                cell.border = Border()
            except AttributeError:
                pass  # phantom merged cell — skip

def _load_csv(path, label):
    if not path.exists():
        print(f"ERROR: {path} not found. Run the backtest first.", file=sys.stderr)
        sys.exit(1)
    df = pd.read_csv(path)
    print(f"Loaded {len(df)} rows from {path.name}  [{label}]")
    return df

def _stats(df):
    fired  = df[df['direction'] != 0].copy()
    priced = fired[
        fired['result'].str.startswith('WIN') | fired['result'].str.startswith('LOSS')
    ]
    wins   = priced[priced['result'].str.startswith('WIN')]
    n_days = len(df)
    n_f    = len(fired)
    n_p    = len(priced)
    n_w    = len(wins)
    wr     = n_w / n_p * 100 if n_p else 0
    pts    = priced['pnl_pts'].sum() if n_p else 0
    inr    = priced['pnl_inr'].sum() if n_p else 0
    best   = priced['pnl_pts'].max() if n_p else 0
    worst  = priced['pnl_pts'].min() if n_p else 0
    d_min  = fired['trade_date'].min() if n_f else df['trade_date'].min()
    d_max  = fired['trade_date'].max() if n_f else df['trade_date'].max()
    da     = (fired['direction'] == fired['actual']).sum() / n_f * 100 if n_f else 0
    return dict(
        fired=fired, priced=priced,
        n_days=n_days, n_f=n_f, n_p=n_p, n_w=n_w,
        wr=wr, pts=pts, inr=inr, best=best, worst=worst,
        d_min=d_min, d_max=d_max, da=da,
        period=f"{d_min[:7]} – {d_max[:7]}",
    )

def _write_trades_sheet(wb, sheet_name, s, lot, index_label):
    ws = wb[sheet_name]
    _clear_sheet(ws)

    COLS = [
        ('Date', 12), ('Dir', 7), ('Strike', 9), ('Side', 6),
        ('Entry', 9), ('Exit', 9), ('PnL pts', 9), ('PnL ₹', 11),
        ('Result', 14), ('Score', 7), ('Exit\nReason', 12), ('Signals', 8),
        ('FII', 6), ('Dir\nAcc', 8),
    ]

    title = (
        f'{index_label} v3 — Backtest Trades  |  {s["period"]}  |  '
        f'Lot: {lot}  |  SL: –50%  TP: +100%'
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    tc = ws.cell(1, 1, title)
    tc.font      = Font(bold=True, color=C_HEADER_FG, name='Arial', size=11)
    tc.fill      = PatternFill('solid', start_color=C_HEADER_BG)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    HDR_ROW = 2
    for ci, (label, width) in enumerate(COLS, start=1):
        _hdr_cell(ws, HDR_ROW, ci, label, width)
    ws.row_dimensions[HDR_ROW].height = 30
    ws.freeze_panes = 'A3'

    DATA_START = 3
    data_rows = s['priced'].sort_values('trade_date').reset_index(drop=True)

    for ri, row in data_rows.iterrows():
        r      = DATA_START + ri
        is_win = str(row.get('result', '')).startswith('WIN')
        bg     = C_WIN_BG if is_win else C_LOSS_BG

        direction = 'LONG' if row['direction'] == 1 else 'SHORT'
        fii_val   = row.get('fii_signature', row.get('fii_fut', ''))
        actual    = row.get('actual', 0)
        dir_ok    = '✓' if row['direction'] == actual else '✗'
        sig_count = int(row.get('signal_count', 0)) if pd.notna(row.get('signal_count')) else ''

        _data_cell(ws, r, 1,  row['trade_date'],                   bg=bg, align='left')
        _data_cell(ws, r, 2,  direction,                            bg=bg)
        _data_cell(ws, r, 3,  row.get('opt_strike', ''),            bg=bg)
        _data_cell(ws, r, 4,  row.get('opt_side', ''),              bg=bg)
        _data_cell(ws, r, 5,  row.get('opt_entry', ''), fmt='0.0',  bg=bg)
        _data_cell(ws, r, 6,  row.get('opt_exit', ''),  fmt='0.0',  bg=bg)
        _data_cell(ws, r, 7,  row.get('pnl_pts', ''),   fmt='+0.0;-0.0;-', bg=bg)
        _data_cell(ws, r, 8,  row.get('pnl_inr', ''),   fmt='₹#,##0;₹-#,##0', bg=bg)
        _data_cell(ws, r, 9,  row.get('result', ''),                bg=bg)
        _data_cell(ws, r, 10, round(row['score'], 3),    fmt='+0.000;-0.000', bg=bg)
        _data_cell(ws, r, 11, row.get('exit_reason', ''),           bg=bg)
        _data_cell(ws, r, 12, sig_count,                            bg=bg)
        _data_cell(ws, r, 13, fii_val,                              bg=bg)
        _data_cell(ws, r, 14, dir_ok,                               bg=bg)
        ws.row_dimensions[r].height = 16

    TOTAL_ROW = DATA_START + len(data_rows)
    ws.merge_cells(start_row=TOTAL_ROW, start_column=1, end_row=TOTAL_ROW, end_column=4)
    _data_cell(ws, TOTAL_ROW, 1, 'TOTAL', bold=True, bg=C_TOTAL_BG, align='center')
    _data_cell(ws, TOTAL_ROW, 5, '', bg=C_TOTAL_BG)
    _data_cell(ws, TOTAL_ROW, 6, '', bg=C_TOTAL_BG)
    _data_cell(ws, TOTAL_ROW, 7,
               f'=SUM(G{DATA_START}:G{TOTAL_ROW-1})',
               fmt='+0.0;-0.0', bold=True, bg=C_TOTAL_BG)
    _data_cell(ws, TOTAL_ROW, 8,
               f'=SUM(H{DATA_START}:H{TOTAL_ROW-1})',
               fmt='₹#,##0;₹-#,##0', bold=True, bg=C_TOTAL_BG)
    _data_cell(ws, TOTAL_ROW, 9,
               f'{s["n_w"]}/{s["n_p"]} wins ({s["wr"]:.1f}%)',
               bold=True, bg=C_TOTAL_BG)
    for c in range(10, len(COLS) + 1):
        _data_cell(ws, TOTAL_ROW, c, '', bg=C_TOTAL_BG)
    ws.row_dimensions[TOTAL_ROW].height = 18

    print(f"{sheet_name}: wrote {len(data_rows)} trades + totals row.")

# ── Load both CSVs ─────────────────────────────────────────────────────────────
df_n  = _load_csv(CSV_N,  'Nifty')
df_bn = _load_csv(CSV_BN, 'BankNifty')

sn  = _stats(df_n)
sbn = _stats(df_bn)

print(f"\nNifty:     {sn['n_p']} trades  WR={sn['wr']:.1f}%  PnL={sn['pts']:+.1f}pts  ₹{sn['inr']:+,.0f}")
print(f"BankNifty: {sbn['n_p']} trades  WR={sbn['wr']:.1f}%  PnL={sbn['pts']:+.1f}pts  ₹{sbn['inr']:+,.0f}")
print(f"Combined:  ₹{sn['inr'] + sbn['inr']:+,.0f}\n")

# ── Open workbook ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX)

# ── Summary sheet ──────────────────────────────────────────────────────────────
ws_sum = wb['Summary']

def _find_row(ws, label):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip().startswith(label):
            return r
    return None

def _set_col(ws, label, col, value):
    r = _find_row(ws, label)
    if r:
        ws.cell(r, col).value = value

combined_inr = sn['inr'] + sbn['inr']
combined_pts = sn['pts'] + sbn['pts']

_set_col(ws_sum, 'Period',          2, f"{sn['d_min'][:10]} – {sn['d_max'][:10]}")
_set_col(ws_sum, 'Data Coverage',   2, f"Sept 2025–{sn['d_max'][:7]} ({sn['n_days']} days)")
_set_col(ws_sum, 'Days Scanned',    2, sn['n_days'])
_set_col(ws_sum, 'Days Traded',     2, sn['n_f'])
_set_col(ws_sum, 'Win Rate',        2, f"{sn['wr']:.1f}%")
_set_col(ws_sum, 'Total PnL (pts)', 2, f"{sn['pts']:+.1f}")
_set_col(ws_sum, 'Total PnL (₹)',   2, f"₹{sn['inr']:+,.0f}")
_set_col(ws_sum, 'Best Trade',      2, f"{sn['best']:+.1f} pts")
_set_col(ws_sum, 'Worst Trade',     2, f"{sn['worst']:+.1f} pts")

# BankNifty column (col 3) — write if rows exist
_set_col(ws_sum, 'Period',          3, f"{sbn['d_min'][:10]} – {sbn['d_max'][:10]}")
_set_col(ws_sum, 'Days Scanned',    3, sbn['n_days'])
_set_col(ws_sum, 'Days Traded',     3, sbn['n_f'])
_set_col(ws_sum, 'Win Rate',        3, f"{sbn['wr']:.1f}%")
_set_col(ws_sum, 'Total PnL (pts)', 3, f"{sbn['pts']:+.1f}")
_set_col(ws_sum, 'Total PnL (₹)',   3, f"₹{sbn['inr']:+,.0f}")
_set_col(ws_sum, 'Best Trade',      3, f"{sbn['best']:+.1f} pts")
_set_col(ws_sum, 'Worst Trade',     3, f"{sbn['worst']:+.1f} pts")

print("Summary sheet updated.")

# ── Trade sheets ───────────────────────────────────────────────────────────────
_write_trades_sheet(wb, 'Nifty v3 Trades',     sn,  LOT_N,  'NIFTY')
_write_trades_sheet(wb, 'BankNifty v3 Trades', sbn, LOT_BN, 'BANKNIFTY')

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(XLSX)
print(f"\nSaved: {XLSX}")
